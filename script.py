from __future__ import annotations

import argparse
import csv
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = " ".join(text.split())
    return text


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def choose_sheet(workbook, requested_sheet: str | None):
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise ValueError(
                f"Feuille introuvable : {requested_sheet}. "
                f"Feuilles disponibles : {', '.join(workbook.sheetnames)}"
            )
        return workbook[requested_sheet]

    # Priorité à la première feuille contenant les colonnes attendues
    for name in workbook.sheetnames:
        ws = workbook[name]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row:
            continue
        headers = {normalize_header(v) for v in first_row if v is not None}
        if {"code", "label fr", "parent immediat"}.issubset(headers):
            return ws

    # Fallback : feuille active
    return workbook[workbook.sheetnames[0]]


def extract_header_map(worksheet) -> dict[str, int]:
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        raise ValueError("Le fichier Excel est vide ou ne contient pas d'en-tête.")

    header_map: dict[str, int] = {}
    for idx, value in enumerate(first_row):
        norm = normalize_header(value)
        if norm:
            header_map[norm] = idx

    required = ["code", "label fr", "parent immediat"]
    missing = [col for col in required if col not in header_map]
    if missing:
        raise ValueError(
            "Colonnes obligatoires introuvables : "
            + ", ".join(missing)
            + f". Colonnes détectées : {', '.join(header_map.keys())}"
        )

    return header_map


def iter_useful_rows(xlsx_path: Path, sheet_name: str | None = None):
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    try:
        ws = choose_sheet(wb, sheet_name)
        header_map = extract_header_map(ws)

        code_idx = header_map["code"]
        label_idx = header_map["label fr"]
        parent_idx = header_map["parent immediat"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            code = clean_cell(row[code_idx] if code_idx < len(row) else None)
            label = clean_cell(row[label_idx] if label_idx < len(row) else None)
            parent = clean_cell(row[parent_idx] if parent_idx < len(row) else None)

            # On ignore les lignes totalement vides ou sans code
            if not code:
                continue

            yield code, label, parent
    finally:
        wb.close()


def write_csv(rows, output_csv: Path) -> int:
    output_csv.parent.mkdir(parents=True, exist_ok=True)

    count = 0
    with output_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["code", "label", "parent"])

        for code, label, parent in rows:
            writer.writerow([code, label, parent])
            count += 1

    return count


def parse_args():
    parser = argparse.ArgumentParser(
        description="Extrait depuis un fichier Excel CIM-10 une table CSV : code;label;parent"
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Chemin du fichier Excel source (.xlsx)",
    )
    parser.add_argument(
        "--output-dir",
        required=True,
        help="Dossier de sortie du CSV",
    )
    parser.add_argument(
        "--output-name",
        default="cim10_hierarchie.csv",
        help="Nom du fichier CSV de sortie (défaut: cim10_hierarchie.csv)",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Nom exact de la feuille à lire. Si absent, détection automatique.",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_csv = output_dir / args.output_name

    if not input_path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {input_path}")

    if input_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Le fichier d'entrée doit être un .xlsx : {input_path}")

    rows = iter_useful_rows(input_path, args.sheet)
    count = write_csv(rows, output_csv)

    print(f"Fichier source : {input_path}")
    print(f"Fichier produit : {output_csv}")
    print(f"Lignes écrites : {count}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Erreur : {e}", file=sys.stderr)
        sys.exit(1)